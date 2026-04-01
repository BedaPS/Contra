"""Node functions for the document processing LangGraph pipeline.

Each node receives DocPipelineState and returns a partial state update dict.
All LLM calls go through LLMAdapter.invoke_vision() — no vendor SDK imports here.
Audit log entry is written BEFORE any state transition (constitution G7).
account_number values are NEVER written to audit log entries (constitution §III).
"""

from __future__ import annotations

import base64
import io
import json
import re
import shutil
import time
from datetime import datetime, timezone
from functools import lru_cache
from pathlib import Path
from statistics import mean
from typing import Any

import fitz  # PyMuPDF
import yaml
from dateutil import parser as dateutil_parser
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from PIL import Image

from src.adapters.llm_adapter import LLMAdapter
from src.audit.logger import AuditEntry
from src.audit.logger import append as audit_append
from src.audit.logger import compute_hash
from src.db.engine import SessionLocal
from src.db.models import BatchRunModel, PaymentRecordModel, RunRecordModel
from src.graph.doc_pipeline.state import DocPipelineState, PaymentRecordDict
from src.settings_store import load_settings

# ── Constants ──
_PROMPT_CONFIGS_DIR = Path(__file__).resolve().parents[3] / "prompt_configs"
_VALID_DOC_TYPES = {"email", "remittance", "receipt", "unknown"}
_MAX_EXTRACTION_ATTEMPTS = 3
_RATE_LIMIT_BACKOFFS = [5, 15, 45]  # seconds: 5 → 15 → 45
_RAW_PIXEL_SIZE_LIMIT = 20 * 1024 * 1024  # 20 MB
_MAX_IMAGE_SIDE = 4096  # pixels
_DEFAULT_THRESHOLD = 0.70

# Excel row fill colours
_FILL_VALID = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_REVIEW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_FAILED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Excel column headers and matching DB attribute names
_XL_HEADERS = [
    "Customer Name", "Account Number", "Payee", "Payment ID", "Payment Method",
    "Payment Date", "Invoice Number", "Reference Doc Number", "Amount Paid",
    "Currency", "Deductions", "Deduction Type", "Notes",
    "Validation Status", "Confidence Score", "LLM Model", "Source File",
]
_XL_ATTRS = [
    "customer_name", "account_number", "payee", "payment_id", "payment_method",
    "payment_date", "invoice_number", "reference_doc_number", "amount_paid",
    "currency", "deductions", "deduction_type", "notes",
    "validation_status", "overall_confidence", "llm_model", "source_filename",
]

# 13 extracted field names (no metadata fields like page_number, confidence_scores, validation_status)
_EXTRACTED_FIELD_NAMES = [
    "customer_name", "account_number", "payee", "payment_id", "payment_method",
    "payment_date", "invoice_number", "reference_doc_number", "amount_paid",
    "currency", "deductions", "deduction_type", "notes",
]

# payment_method normalisation table (lowercase key → canonical value)
_PAYMENT_METHOD_MAP: dict[str, str] = {
    "eft": "EFT",
    "electronic funds transfer": "EFT",
    "electronic transfer": "EFT",
    "electronic fund transfer": "EFT",
    "cash": "CASH",
    "cheque": "CHEQUE",
    "check": "CHEQUE",
    "chq": "CHEQUE",
    "chk": "CHEQUE",
    "direct deposit": "DIRECT_DEPOSIT",
    "direct_deposit": "DIRECT_DEPOSIT",
    "dd": "DIRECT_DEPOSIT",
    "card": "CARD",
    "credit card": "CARD",
    "debit card": "CARD",
    "cc": "CARD",
    "dc": "CARD",
}

# currency symbol / name → ISO 4217 mapping
_CURRENCY_MAP: dict[str, str] = {
    "$": "USD",
    "usd": "USD",
    "us dollar": "USD",
    "€": "EUR",
    "eur": "EUR",
    "euro": "EUR",
    "£": "GBP",
    "gbp": "GBP",
    "pound": "GBP",
    "sterling": "GBP",
    "r": "ZAR",
    "zar": "ZAR",
    "rand": "ZAR",
    "south african rand": "ZAR",
    "a$": "AUD",
    "aud": "AUD",
    "aus dollar": "AUD",
    "c$": "CAD",
    "cad": "CAD",
    "¥": "JPY",
    "jpy": "JPY",
    "yen": "JPY",
    "cny": "CNY",
    "rmb": "CNY",
    "yuan": "CNY",
    "chf": "CHF",
    "inr": "INR",
    "₹": "INR",
    "sgd": "SGD",
    "hkd": "HKD",
    "nzd": "NZD",
    "myr": "MYR",
    "ngn": "NGN",
    "kes": "KES",
    "ghs": "GHS",
}

# Classifier prompt is doc-type-agnostic — embedded here, not in YAML
_CLASSIFIER_PROMPT = (
    "Classify this payment document. Respond with EXACTLY one word from this list:\n"
    "  email       - payment notification email or remittance email message\n"
    "  remittance  - remittance advice slip, statement, or printed remittance document\n"
    "  receipt     - payment receipt, proof of payment, or till/cash slip\n"
    "  unknown     - cannot be classified as any of the above\n\n"
    "Return ONLY one word. No explanation, no punctuation."
)


# ── Private helpers ──

@lru_cache(maxsize=4)
def _load_prompt_config(doc_type: str) -> dict:
    """Load and cache a YAML prompt config. Called once per doc_type at runtime."""
    config_path = _PROMPT_CONFIGS_DIR / f"{doc_type}.yaml"
    with open(config_path, encoding="utf-8") as fh:
        return yaml.safe_load(fh)


def _render_page_to_base64(page: fitz.Page) -> str:
    """Render a PDF page to a base64-encoded PNG at 150 DPI.

    If the raw pixel buffer exceeds 20 MB, the image is resized to ≤ 4096 px
    on the longest side before encoding.
    """
    mat = fitz.Matrix(150 / 72, 150 / 72)  # 150 DPI
    pix = page.get_pixmap(matrix=mat, alpha=False)

    if len(pix.samples) > _RAW_PIXEL_SIZE_LIMIT:
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        max_side = max(img.width, img.height)
        if max_side > _MAX_IMAGE_SIDE:
            scale = _MAX_IMAGE_SIDE / max_side
            img = img.resize(
                (int(img.width * scale), int(img.height * scale)),
                Image.LANCZOS,
            )
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        png_bytes = buf.getvalue()
    else:
        png_bytes = pix.tobytes("png")

    return base64.b64encode(png_bytes).decode()


def _is_rate_limit_error(exc: Exception) -> bool:
    """Detect a 429 / rate-limit error without importing vendor SDKs."""
    type_name = type(exc).__name__.lower()
    msg = str(exc).lower()
    return (
        "ratelimit" in type_name
        or "rate_limit" in type_name
        or "429" in msg
        or "rate limit" in msg
    )


def _strip_json_fences(text: str) -> str:
    """Remove markdown code fences that vision LLMs sometimes prepend."""
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```[a-z]*\n?", "", text)
        text = re.sub(r"\n?```$", "", text)
    return text.strip()


def _normalise_amount(val: Any) -> float | None:
    """Convert a raw amount value (string, int, float, or None) to float."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    cleaned = re.sub(r"[^\d.\-]", "", str(val).replace(",", ""))
    try:
        return float(cleaned)
    except ValueError:
        return None


def _normalise_date(val: str | None) -> str | None:
    """Normalise any date string to YYYY-MM-DD. Returns original if unparseable."""
    if not val:
        return None
    try:
        dt = dateutil_parser.parse(str(val), dayfirst=True)
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return val


def _normalise_currency(val: str | None) -> str | None:
    """Map currency symbols/names to ISO 4217 codes. Passes through 3-letter codes."""
    if not val:
        return None
    key = str(val).strip().lower()
    if key in _CURRENCY_MAP:
        return _CURRENCY_MAP[key]
    upper = str(val).strip().upper()
    if re.match(r"^[A-Z]{3}$", upper):
        return upper  # Already a valid ISO 4217 code
    return val


def _normalise_payment_method(val: str | None) -> str | None:
    """Map payment method variants to canonical form (EFT/CASH/CHEQUE/DIRECT_DEPOSIT/CARD)."""
    if not val:
        return None
    key = str(val).strip().lower()
    return _PAYMENT_METHOD_MAP.get(key, str(val).upper())


def _compute_overall_confidence(confidence_scores: dict[str, float]) -> float:
    """Return the mean of all confidence score values, or 0.0 if empty."""
    if not confidence_scores:
        return 0.0
    return mean(confidence_scores.values())


def _redacted_confidence_summary(records: list[PaymentRecordDict]) -> dict[str, float]:
    """Aggregate per-field confidence scores into a summary dict — no PII values."""
    agg: dict[str, list[float]] = {}
    for rec in records:
        for field, score in (rec.get("confidence_scores") or {}).items():
            agg.setdefault(field, []).append(score)
    return {field: mean(scores) for field, scores in agg.items()}


def _is_null_amount(val: Any) -> bool:
    """Return True if a value represents a null / missing amount."""
    if val is None:
        return True
    if isinstance(val, str) and (not val.strip() or val.strip().lower() in ("null", "none", "")):
        return True
    return False


def _all_null(records: list[PaymentRecordDict]) -> bool:
    """Return True if every record's amount_paid is effectively null."""
    return all(_is_null_amount(r.get("amount_paid")) for r in records)


def _build_extraction_prompt(prompt_config: dict, page_number: int) -> str:
    """Build the per-page extraction prompt from the doc-type YAML config."""
    context_hint = prompt_config.get("context_hint", "")
    field_hints = prompt_config.get("field_hints") or {}
    field_hints_text = "\n".join(
        f"  - {field}: {hint}" for field, hint in field_hints.items()
    )
    return (
        f"{context_hint}\n\n"
        "Extract ALL payment records visible on this page. "
        "For each distinct payment amount found, create a separate record.\n\n"
        "Fields to extract:\n"
        f"{field_hints_text}\n\n"
        "Return a JSON array. Each item MUST include:\n"
        "- All 13 payment fields (use null for any missing field):\n"
        "  customer_name, account_number, payee, payment_id, payment_method,\n"
        "  payment_date, invoice_number, reference_doc_number, amount_paid,\n"
        "  currency, deductions, deduction_type, notes\n"
        f'- "page_number": {page_number}  (1-based page index)\n'
        '- "confidence_scores": {"field_name": 0.0-1.0} for each of the 13 fields\n\n'
        "Return ONLY the JSON array. No prose, no markdown code fence."
    )


def _parse_extracted_records(raw: str, page_number: int) -> list[PaymentRecordDict] | None:
    """Parse a raw LLM string response into a list of PaymentRecordDict.

    Returns None if JSON parsing fails or no valid records are found.
    account_number values are stored as-is; they will be written to DB only,
    never to audit logs.
    """
    text = _strip_json_fences(raw)
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        return None

    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list):
        return None

    records: list[PaymentRecordDict] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        record: PaymentRecordDict = {
            "customer_name": item.get("customer_name"),
            "account_number": item.get("account_number"),
            "payee": item.get("payee"),
            "payment_id": item.get("payment_id"),
            "payment_method": item.get("payment_method"),
            "payment_date": item.get("payment_date"),
            "invoice_number": item.get("invoice_number"),
            "reference_doc_number": item.get("reference_doc_number"),
            "amount_paid": item.get("amount_paid"),   # normaliser_node handles conversion
            "currency": item.get("currency"),
            "deductions": item.get("deductions"),     # normaliser_node handles conversion
            "deduction_type": item.get("deduction_type"),
            "notes": item.get("notes"),
            "page_number": item.get("page_number", page_number),
            "confidence_scores": item.get("confidence_scores") or {},
        }
        records.append(record)

    return records if records else None


# ── Node functions ──

def classifier_node(state: DocPipelineState) -> dict[str, Any]:
    """Render page 1 of the work file → vision LLM classifies → sets doc_type.

    Loads and caches the YAML prompt config for the classified doc_type.
    Writes audit log entry. Routes to error_node on any failure.
    """
    run_record_id = state.get("run_record_id", "unknown")
    work_file_path = state.get("work_file_path", "")
    source_file_path = state.get("source_file_path", "")

    # Render page 1
    try:
        doc = fitz.open(work_file_path)
        page_b64 = _render_page_to_base64(doc[0])
        doc.close()
    except Exception as exc:
        audit_append(AuditEntry(
            agent="classifier_node",
            input_hash=compute_hash({"run_record_id": run_record_id, "file": work_file_path}),
            output_hash=compute_hash({"error": str(exc)}),
            state_from="Ingested",
            state_to="Error",
            decision="RENDER_FAILED",
            rationale=f"Failed to render page 1 for classification: {type(exc).__name__}: {exc}",
        ))
        return {"error": f"Page render failed: {exc}", "error_type": "render_error"}

    # Classify via LLM
    try:
        adapter = LLMAdapter()
        raw_response = adapter.invoke_vision(_CLASSIFIER_PROMPT, [page_b64])
        doc_type = raw_response.strip().lower().split()[0] if raw_response.strip() else "unknown"
        if doc_type not in _VALID_DOC_TYPES:
            doc_type = "unknown"
    except Exception as exc:
        audit_append(AuditEntry(
            agent="classifier_node",
            input_hash=compute_hash({"run_record_id": run_record_id}),
            output_hash=compute_hash({"error": str(exc)}),
            state_from="Ingested",
            state_to="Error",
            decision="LLM_FAILED",
            rationale=f"Vision LLM classify call failed: {type(exc).__name__}: {exc}",
        ))
        return {"error": f"Classifier LLM call failed: {exc}", "error_type": "llm_error"}

    # Load prompt config (cached)
    try:
        prompt_config: dict = dict(_load_prompt_config(doc_type))
    except Exception:
        prompt_config = {}

    audit_append(AuditEntry(
        agent="classifier_node",
        input_hash=compute_hash({
            "run_record_id": run_record_id,
            "filename": Path(source_file_path).name,
        }),
        output_hash=compute_hash({"doc_type": doc_type}),
        state_from="Ingested",
        state_to="Classified",
        decision="CLASSIFIED",
        rationale=f"Document classified as '{doc_type}'.",
    ))

    return {
        "doc_type": doc_type,
        "prompt_config": prompt_config,
        "error": None,
        "error_type": None,
    }


def extractor_node(state: DocPipelineState) -> dict[str, Any]:
    """Render all pages → per-page vision LLM extraction → list[PaymentRecordDict].

    Retries up to 3 times total on JSON parse failure. Applies exponential
    backoff (5 s → 15 s → 45 s) on HTTP 429 / rate-limit errors.
    Routes to error_node if all records have null amount_paid or 3 parse failures.
    """
    run_record_id = state.get("run_record_id", "unknown")
    work_file_path = state.get("work_file_path", "")
    doc_type = state.get("doc_type", "unknown")
    prompt_config = state.get("prompt_config") or {}
    if not prompt_config:
        try:
            prompt_config = dict(_load_prompt_config(doc_type))
        except Exception:
            prompt_config = {}

    # Render all pages
    try:
        doc = fitz.open(work_file_path)
        page_images = [_render_page_to_base64(page) for page in doc]
        doc.close()
    except Exception as exc:
        audit_append(AuditEntry(
            agent="extractor_node",
            input_hash=compute_hash({"run_record_id": run_record_id}),
            output_hash=compute_hash({"error": str(exc)}),
            state_from="Classified",
            state_to="Error",
            decision="RENDER_FAILED",
            rationale=f"Failed to render pages for extraction: {type(exc).__name__}: {exc}",
        ))
        return {
            "error": f"Page render failed: {exc}",
            "error_type": "render_error",
            "extraction_attempts": state.get("extraction_attempts", 0) + 1,
        }

    adapter = LLMAdapter()
    all_records: list[PaymentRecordDict] = []
    total_attempts = 0

    for page_idx, page_b64 in enumerate(page_images, start=1):
        prompt = _build_extraction_prompt(prompt_config, page_idx)
        page_records: list[PaymentRecordDict] | None = None
        page_attempt = 0

        while page_records is None and page_attempt < _MAX_EXTRACTION_ATTEMPTS:
            page_attempt += 1
            total_attempts += 1
            try:
                raw = adapter.invoke_vision(prompt, [page_b64])
                page_records = _parse_extracted_records(raw, page_idx)
                if page_records is None and page_attempt < _MAX_EXTRACTION_ATTEMPTS:
                    continue  # retry on parse failure
            except Exception as exc:
                if _is_rate_limit_error(exc):
                    backoff_idx = min(page_attempt - 1, len(_RATE_LIMIT_BACKOFFS) - 1)
                    time.sleep(_RATE_LIMIT_BACKOFFS[backoff_idx])
                    if page_attempt >= _MAX_EXTRACTION_ATTEMPTS:
                        audit_append(AuditEntry(
                            agent="extractor_node",
                            input_hash=compute_hash({"run_record_id": run_record_id, "page": page_idx}),
                            output_hash=compute_hash({"error": "rate_limit_exhausted"}),
                            state_from="Classified",
                            state_to="Error",
                            decision="RATE_LIMIT_EXHAUSTED",
                            rationale=(
                                f"Rate limit on page {page_idx} after {page_attempt} attempt(s). "
                                f"Backoff sequence applied: {_RATE_LIMIT_BACKOFFS}."
                            ),
                        ))
                        return {
                            "error": f"Rate limit exceeded on page {page_idx} after {page_attempt} attempts.",
                            "error_type": "rate_limit",
                            "extraction_attempts": total_attempts,
                        }
                    continue  # retry after backoff
                else:
                    audit_append(AuditEntry(
                        agent="extractor_node",
                        input_hash=compute_hash({"run_record_id": run_record_id, "page": page_idx}),
                        output_hash=compute_hash({"error": str(exc)}),
                        state_from="Classified",
                        state_to="Error",
                        decision="LLM_FAILED",
                        rationale=(
                            f"LLM extraction failed on page {page_idx}, "
                            f"attempt {page_attempt}: {type(exc).__name__}"
                        ),
                    ))
                    return {
                        "error": f"LLM extraction failed on page {page_idx}: {exc}",
                        "error_type": "llm_error",
                        "extraction_attempts": total_attempts,
                    }

        if page_records is None:
            # All parse attempts exhausted for this page
            audit_append(AuditEntry(
                agent="extractor_node",
                input_hash=compute_hash({"run_record_id": run_record_id, "page": page_idx}),
                output_hash=compute_hash({"error": "parse_failure"}),
                state_from="Classified",
                state_to="Error",
                decision="PARSE_FAILED",
                rationale=(
                    f"JSON parse failed for page {page_idx} "
                    f"after {page_attempt} attempt(s)."
                ),
            ))
            return {
                "error": (
                    f"Failed to parse extraction response for page {page_idx} "
                    f"after {page_attempt} attempts."
                ),
                "error_type": "parse_error",
                "extraction_attempts": total_attempts,
            }

        all_records.extend(page_records)

    # Check for all-null result
    if not all_records or _all_null(all_records):
        audit_append(AuditEntry(
            agent="extractor_node",
            input_hash=compute_hash({"run_record_id": run_record_id}),
            output_hash=compute_hash({"records": len(all_records)}),
            state_from="Classified",
            state_to="Error",
            decision="ALL_NULL",
            rationale=(
                f"Extraction produced {len(all_records)} record(s) "
                "but all amount_paid values are null."
            ),
        ))
        return {
            "error": "Extraction produced no usable records (all amount_paid values are null).",
            "error_type": "all_null",
            "raw_records": all_records,
            "extraction_attempts": total_attempts,
        }

    confidence_summary = _redacted_confidence_summary(all_records)
    audit_append(AuditEntry(
        agent="extractor_node",
        input_hash=compute_hash({"run_record_id": run_record_id, "pages": len(page_images)}),
        output_hash=compute_hash({"records": len(all_records)}),
        state_from="Classified",
        state_to="Extracted",
        decision="EXTRACTED",
        rationale=(
            f"Extracted {len(all_records)} record(s) across "
            f"{len(page_images)} page(s) in {total_attempts} attempt(s)."
        ),
        confidence_scores=confidence_summary,
    ))

    return {
        "raw_records": all_records,
        "page_images": page_images,
        "extraction_attempts": total_attempts,
        "error": None,
        "error_type": None,
    }


def normaliser_node(state: DocPipelineState) -> dict[str, Any]:
    """Normalise raw extracted records — no LLM calls.

    Normalisation rules:
    - payment_date → YYYY-MM-DD (dateutil.parser, dayfirst=True)
    - amount_paid / deductions → float (strips symbols, commas)
    - currency → ISO 4217 code
    - payment_method → EFT | CASH | CHEQUE | DIRECT_DEPOSIT | CARD
    """
    run_record_id = state.get("run_record_id", "unknown")
    raw_records = state.get("raw_records") or []
    normalised: list[PaymentRecordDict] = []

    for rec in raw_records:
        norm = dict(rec)
        norm["payment_date"] = _normalise_date(rec.get("payment_date"))
        norm["amount_paid"] = _normalise_amount(rec.get("amount_paid"))
        norm["deductions"] = _normalise_amount(rec.get("deductions"))
        norm["currency"] = _normalise_currency(rec.get("currency"))
        norm["payment_method"] = _normalise_payment_method(rec.get("payment_method"))
        normalised.append(norm)  # type: ignore[arg-type]

    audit_append(AuditEntry(
        agent="normaliser_node",
        input_hash=compute_hash({"run_record_id": run_record_id, "raw_count": len(raw_records)}),
        output_hash=compute_hash({"normalised_count": len(normalised)}),
        state_from="Extracted",
        state_to="Normalised",
        decision="NORMALISED",
        rationale=(
            f"Normalised {len(normalised)} record(s): "
            "dates→YYYY-MM-DD, amounts→float, currency→ISO4217, "
            "payment_method→canonical."
        ),
    ))

    return {
        "normalised_records": normalised,
        "error": None,
        "error_type": None,
    }


def validator_node(state: DocPipelineState) -> dict[str, Any]:
    """Validate each normalised record against per-field YAML confidence thresholds.

    Three-rule priority (FR-009):
    1. amount_paid is None OR confidence < threshold → 'Extraction Failed'
    2. Any other field confidence < its threshold  → 'Review Required'
    3. All fields meet thresholds                  → 'Valid'

    Writes field names and confidence scores to audit — no PII values.
    """
    run_record_id = state.get("run_record_id", "unknown")
    normalised_records = state.get("normalised_records") or []
    prompt_config = state.get("prompt_config") or {}
    thresholds: dict[str, float] = prompt_config.get("confidence_thresholds") or {}

    validated: list[PaymentRecordDict] = []

    for rec in normalised_records:
        confidence_scores = rec.get("confidence_scores") or {}
        amount_threshold = thresholds.get("amount_paid", 0.90)
        amount_confidence = confidence_scores.get("amount_paid", 0.0)

        # Rule 1: amount_paid missing or below threshold
        if rec.get("amount_paid") is None or amount_confidence < amount_threshold:
            status = "Extraction Failed"
        else:
            # Rule 2: any other field below its threshold
            below_threshold = any(
                confidence_scores.get(field, 1.0) < thresholds.get(field, _DEFAULT_THRESHOLD)
                for field in confidence_scores
                if field != "amount_paid"
            )
            status = "Review Required" if below_threshold else "Valid"

        validated_rec = dict(rec)
        validated_rec["validation_status"] = status
        validated.append(validated_rec)  # type: ignore[arg-type]

    # Build status summary and confidence summary (no PII values)
    status_counts: dict[str, int] = {}
    for rec in validated:
        s = rec.get("validation_status", "")
        status_counts[s] = status_counts.get(s, 0) + 1

    confidence_summary = _redacted_confidence_summary(validated)

    audit_append(AuditEntry(
        agent="validator_node",
        input_hash=compute_hash({"run_record_id": run_record_id, "count": len(normalised_records)}),
        output_hash=compute_hash({"status_counts": status_counts}),
        state_from="Normalised",
        state_to="Validated",
        decision="VALIDATED",
        rationale=f"Validated {len(validated)} record(s): {status_counts}.",
        confidence_scores=confidence_summary,
    ))

    return {
        "validated_records": validated,
        "error": None,
        "error_type": None,
    }


def excel_writer_node(state: DocPipelineState) -> dict[str, Any]:
    """Persist validated records to DB, update run metadata, write Excel and accuracy log.

    AUDIT LOG IS WRITTEN FIRST — constitution G7 (before any DB or file write).
    account_number values are written to the DB but NEVER to audit log entries.
    """
    run_record_id = state.get("run_record_id", "unknown")
    batch_id = state.get("batch_id", "unknown")
    validated_records = state.get("validated_records") or []
    doc_type = state.get("doc_type", "unknown")
    extraction_attempts = state.get("extraction_attempts", 1)
    source_filename = Path(state.get("source_file_path", "unknown")).name

    settings = load_settings()

    # ── AUDIT FIRST (constitution G7) ──
    confidence_summary = _redacted_confidence_summary(validated_records)
    overall_by_record = [
        _compute_overall_confidence(rec.get("confidence_scores") or {})
        for rec in validated_records
    ]
    overall_confidence = mean(overall_by_record) if overall_by_record else 0.0
    status_counts: dict[str, int] = {}
    for rec in validated_records:
        s = rec.get("validation_status", "")
        status_counts[s] = status_counts.get(s, 0) + 1

    audit_append(AuditEntry(
        agent="excel_writer_node",
        input_hash=compute_hash({
            "run_record_id": run_record_id,
            "batch_id": batch_id,
            "record_count": len(validated_records),
        }),
        output_hash=compute_hash({
            "status_counts": status_counts,
            "overall_confidence": round(overall_confidence, 4),
        }),
        state_from="Validated",
        state_to="Finalised",
        decision="WRITE",
        rationale=(
            f"Writing {len(validated_records)} record(s) for "
            f"run_record_id={run_record_id}. "
            f"Status breakdown: {status_counts}. "
            f"Overall confidence: {overall_confidence:.4f}."
        ),
        confidence_scores=confidence_summary,
    ))

    # ── DB: batch-insert PaymentRecords ──
    now_utc = datetime.now(timezone.utc)
    with SessionLocal() as session:
        for rec in validated_records:
            scores = rec.get("confidence_scores") or {}
            rec_overall = _compute_overall_confidence(scores)
            db_record = PaymentRecordModel(
                run_record_id=run_record_id,
                batch_id=batch_id,
                page_number=rec.get("page_number"),
                customer_name=rec.get("customer_name"),
                account_number=rec.get("account_number"),  # stored in DB; never in audit
                payee=rec.get("payee"),
                payment_id=rec.get("payment_id"),
                payment_method=rec.get("payment_method"),
                payment_date=rec.get("payment_date"),
                invoice_number=rec.get("invoice_number"),
                reference_doc_number=rec.get("reference_doc_number"),
                amount_paid=rec.get("amount_paid"),
                currency=rec.get("currency"),
                deductions=rec.get("deductions"),
                deduction_type=rec.get("deduction_type"),
                notes=rec.get("notes"),
                validation_status=rec.get("validation_status", "Extraction Failed"),
                confidence_scores=json.dumps(scores),
                overall_confidence=rec_overall,
                llm_provider=settings.provider,
                llm_model=settings.model,
                source_filename=source_filename,
                doc_type=doc_type,
                created_at=now_utc,
            )
            session.add(db_record)

        # Update RunRecord → Completed
        run_rec = session.get(RunRecordModel, run_record_id)
        if run_rec:
            run_rec.status = "Completed"
            run_rec.record_count = len(validated_records)
            run_rec.completed_at = now_utc

        # Update BatchRun total_records
        batch_run = session.get(BatchRunModel, batch_id)
        if batch_run:
            batch_run.total_records = (batch_run.total_records or 0) + len(validated_records)

        session.commit()

    # ── accuracy.jsonl ──
    output_dir_str = settings.output_directory
    if output_dir_str:
        out_path = Path(output_dir_str)
        out_path.mkdir(parents=True, exist_ok=True)

        # Overall doc-level validation_status — most critical wins
        if status_counts.get("Extraction Failed", 0) > 0:
            doc_validation_status = "Extraction Failed"
        elif status_counts.get("Review Required", 0) > 0:
            doc_validation_status = "Review Required"
        else:
            doc_validation_status = "Valid"

        total_possible = len(_EXTRACTED_FIELD_NAMES) * len(validated_records)
        null_fields = sum(
            1
            for rec in validated_records
            for f in _EXTRACTED_FIELD_NAMES
            if rec.get(f) is None
        )
        fields_extracted = total_possible - null_fields

        accuracy_entry = {
            "timestamp": now_utc.isoformat(),
            "source_filename": source_filename,
            "doc_type": doc_type,
            "validation_status": doc_validation_status,
            "overall_confidence": round(overall_confidence, 4),
            "fields_extracted": fields_extracted,
            "null_fields": null_fields,
            "extraction_attempts": extraction_attempts,
            "llm_provider": settings.provider,
            "llm_model": settings.model,
        }
        jsonl_path = out_path / "accuracy.jsonl"
        with open(jsonl_path, "a", encoding="utf-8") as jf:
            jf.write(json.dumps(accuracy_entry) + "\n")

        # ── results.xlsx: rewrite from DB ──
        _rewrite_results_xlsx(out_path)

    return {
        "error": None,
        "error_type": None,
    }


def _rewrite_results_xlsx(output_dir: Path) -> None:
    """Query all PaymentRecords from DB and rewrite results.xlsx with two sheets."""
    with SessionLocal() as session:
        all_records = session.query(PaymentRecordModel).all()

    wb = Workbook()

    # Sheet 1: All records
    ws1 = wb.active
    ws1.title = "Payment Records"
    ws1.append(_XL_HEADERS)

    # Sheet 2: Review Required + Extraction Failed only
    ws2 = wb.create_sheet("Review Required")
    ws2.append(_XL_HEADERS)

    for record in all_records:
        row = [getattr(record, attr, None) for attr in _XL_ATTRS]
        status = record.validation_status or ""

        if status == "Valid":
            fill = _FILL_VALID
        elif status == "Review Required":
            fill = _FILL_REVIEW
        else:
            fill = _FILL_FAILED

        ws1.append(row)
        for col_idx in range(1, len(_XL_HEADERS) + 1):
            ws1.cell(row=ws1.max_row, column=col_idx).fill = fill

        if status in ("Review Required", "Extraction Failed"):
            ws2.append(row)
            for col_idx in range(1, len(_XL_HEADERS) + 1):
                ws2.cell(row=ws2.max_row, column=col_idx).fill = fill

    xlsx_path = output_dir / "results.xlsx"
    wb.save(str(xlsx_path))


def error_node(state: DocPipelineState) -> dict[str, Any]:
    """Handle pipeline errors: write audit, mark RunRecord as Failed, move work file.

    Constitution G7: audit log is written before any file system operations.
    """
    run_record_id = state.get("run_record_id", "unknown")
    error_msg = state.get("error", "Unknown error")
    error_type = state.get("error_type", "unknown")
    work_file_path = state.get("work_file_path", "")

    # ── AUDIT FIRST (constitution G7) ──
    audit_append(AuditEntry(
        agent="error_node",
        input_hash=compute_hash({"run_record_id": run_record_id, "error_type": error_type}),
        output_hash=compute_hash({"status": "Failed"}),
        state_from="Error",
        state_to="Failed",
        decision="FAILED",
        rationale=(
            f"Pipeline failed for run_record_id={run_record_id}. "
            f"error_type={error_type}. "
            f"Message: {str(error_msg)[:200]}"
        ),
    ))

    # Update RunRecord status → Failed
    now_utc = datetime.now(timezone.utc)
    with SessionLocal() as session:
        run_rec = session.get(RunRecordModel, run_record_id)
        if run_rec:
            run_rec.status = "Failed"
            run_rec.completed_at = now_utc
            session.commit()

    # Move work file to failed/ subdirectory
    if work_file_path:
        try:
            src = Path(work_file_path)
            if src.exists():
                failed_dir = src.parent / "failed"
                failed_dir.mkdir(parents=True, exist_ok=True)
                shutil.move(str(src), str(failed_dir / src.name))
        except OSError:
            pass  # Non-fatal: file-move errors should not obscure the original error

    return {
        "error": error_msg,
        "error_type": error_type,
    }
